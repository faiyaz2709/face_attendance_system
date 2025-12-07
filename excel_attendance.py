# excel_attendance.py
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

EXCEL_DIR = "attendance"
EXCEL_FILE = os.path.join(EXCEL_DIR, "attendance_data.xlsx")
SHEET_NAME = "Attendance"

def create_file_if_missing():
    os.makedirs(EXCEL_DIR, exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["Name", "Date", "Time", "Status"])  # header
        wb.save(EXCEL_FILE)
        print(f"Created new Excel file: {EXCEL_FILE}")

def already_marked_today(name):
    """Return True if name already has a row with today's date."""
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    today = datetime.now().strftime("%Y-%m-%d")
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        r_name, r_date, r_time, r_status = row
        if r_name == name and r_date == today:
            return True
    return False

def mark_attendance(name, status="Present"):
    create_file_if_missing()
    if already_marked_today(name):
        print(f"[Excel] Already marked today: {name}")
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    now = datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")
    ws.append([name, date, time, status])
    wb.save(EXCEL_FILE)
    print(f"[Excel] Marked attendance: {name} | {date} {time}")
    return True

# test when run directly
if __name__ == "__main__":
    create_file_if_missing()
    n = input("Name to mark: ")
    mark_attendance(n)
